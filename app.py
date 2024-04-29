from flask import Flask, render_template, request
from openpyxl import load_workbook
import requests
from lxml import html
import urllib3
from urllib3.exceptions import InsecureRequestWarning

# Desativar os avisos de certificado SSL
urllib3.disable_warnings(InsecureRequestWarning)


app = Flask(__name__)

def carregar_dados_planilha():
    wb = load_workbook('codigo_rastreio.xlsx', data_only=True)
    sheet = wb['Planilha1']
    dados = {}
    for row in sheet.iter_rows(values_only=True):
        codigo = row[0]  # O código está na primeira coluna (coluna A)
        placa = row[1]   # A placa está na segunda coluna (coluna B)
        if codigo and placa:
            dados[codigo] = placa
    return dados

# Função para buscar informações de rastreamento para um código específico diretamente no site dos Correios
def buscar_informacoes_rastreamento(codigo):
    url = f'https://www.sitecorreios.com.br/{codigo}'  # Substitua 'linkdoseucodigo' pelo link correto
    response = requests.get(url, verify=False)
    if response.status_code == 200:
        tree = html.fromstring(response.content)
        informacoes = tree.xpath("//div[contains(@class, 'relative pb-10') or contains(@class, 'ml-5 flex flex-col mt-2')]")
        if informacoes:
            return [info.text_content().strip() for info in informacoes]
        else:
            return ["Informações não encontradas"]
    else:
        return ["Informações não encontradas"]

@app.route('/', methods=['GET', 'POST'])
def rastreios():
    codigos = carregar_dados_planilha()
    informacoes = {}
    for codigo in codigos:
        informacoes[codigo] = buscar_informacoes_rastreamento(codigo)
    termo_pesquisa = ''
    tipo_pesquisa = 'placa'  # Valor padrão para o tipo de pesquisa
    if request.method == 'POST':
        termo_pesquisa = request.form['termo_pesquisa']
        tipo_pesquisa = request.form['tipo_pesquisa']
        informacoes_filtradas = {}
        mensagem = None
        if tipo_pesquisa == 'placa':
            informacoes_filtradas = {c: info for c, info in informacoes.items() if codigos[c] == termo_pesquisa.upper()}
            if not informacoes_filtradas:
                mensagem = f"Não existem dados para a placa {termo_pesquisa} na planilha."
        elif tipo_pesquisa == 'codigo':
            termos_pesquisa = [termo.strip() for termo in termo_pesquisa.split(',')]
            for termo in termos_pesquisa:
                if termo in codigos:
                    informacoes_filtradas[termo] = informacoes[termo]
                    if not informacoes_filtradas[termo]:
                        mensagem = f"Não foi possível encontrar informações para o código de rastreamento {termo}."
                else:
                    informacoes_filtradas[termo] = buscar_informacoes_rastreamento(termo)
                    if not informacoes_filtradas[termo]:
                        mensagem = f"Não foi possível encontrar informações para o código de rastreamento {termo}."
        return render_template('rastreamento.html', informacoes=informacoes_filtradas, termo_pesquisa=termo_pesquisa, tipo_pesquisa=tipo_pesquisa, codigos=codigos, mensagem=mensagem)
    elif request.method == 'GET':
        return render_template('rastreamento.html', informacoes=informacoes, termo_pesquisa=termo_pesquisa, tipo_pesquisa=tipo_pesquisa, codigos=codigos)

if __name__ == '__main__':
    app.run(debug=True)