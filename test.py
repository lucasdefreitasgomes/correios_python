from flask import Flask, render_template, request
from openpyxl import load_workbook
import requests
from lxml import html
import pandas as pd
import locale
import logging

logging.basicConfig(level=logging.DEBUG)

app = Flask(__name__)

def carregar_dados_planilha():
    wb = load_workbook('codigo_rastreio.xlsx', data_only=True)
    sheet = wb['Planilha1']
    dados = {}
    for row in sheet.iter_rows(values_only=True):
        codigo = row[0]  # O código está na primeira coluna (coluna A)
        placa = row[1]   # A placa está na segunda coluna (coluna B)
        if codigo and placa:
            dados[codigo] = placa
    return dados

# Função para buscar informações de rastreamento para um código específico diretamente no site dos Correios
def buscar_informacoes_rastreamento(codigo):
    url = f'https://www.sitecorreios.com.br/{codigo}'  # Substitua 'linkdoseucodigo' pelo link correto
    response = requests.get(url)
    if response.status_code == 200:
        tree = html.fromstring(response.content)
        informacoes = tree.xpath("//div[contains(@class, 'relative pb-10') or contains(@class, 'ml-5 flex flex-col mt-2')]")
        if informacoes:
            return [info.text_content().strip() for info in informacoes]
        else:
            return ["Informações não encontradas"]
    else:
        return ["Informações não encontradas"]

@app.route('/rastreios', methods=['GET', 'POST'])
def rastreios():
    codigos = carregar_dados_planilha()
    logging.debug(f"Códigos carregados: {codigos}")
    informacoes = {}
    for codigo in codigos:
        try:
            informacoes[codigo] = buscar_informacoes_rastreamento(codigo)
            logging.debug(f"Informações para código {codigo}: {informacoes[codigo]}")
        except Exception as e:
            logging.error(f"Erro ao buscar informações para código {codigo}: {str(e)}")
            informacoes[codigo] = ["Erro ao buscar informações"]
    termo_pesquisa = ''
    tipo_pesquisa = 'placa'  # Valor padrão para o tipo de pesquisa
    if request.method == 'POST':
        termo_pesquisa = request.form['termo_pesquisa']
        tipo_pesquisa = request.form['tipo_pesquisa']
        informacoes_filtradas = {}
        mensagem = None
        if tipo_pesquisa == 'placa':
            informacoes_filtradas = {c: info for c, info in informacoes.items() if codigos[c] == termo_pesquisa.upper()}
            if not informacoes_filtradas:
                mensagem = f"Não existem dados para a placa {termo_pesquisa} na planilha."
        elif tipo_pesquisa == 'codigo':
            termos_pesquisa = [termo.strip() for termo in termo_pesquisa.split(',')]
            for termo in termos_pesquisa:
                if termo in codigos:
                    informacoes_filtradas[termo] = informacoes[termo]
                    if not informacoes_filtradas[termo]:
                        mensagem = f"Não foi possível encontrar informações para o código de rastreamento {termo}."
                else:
                    informacoes_filtradas[termo] = buscar_informacoes_rastreamento(termo)
                    if not informacoes_filtradas[termo]:
                        mensagem = f"Não foi possível encontrar informações para o código de rastreamento {termo}."
        return render_template('rastreamento.html', informacoes=informacoes_filtradas, termo_pesquisa=termo_pesquisa, tipo_pesquisa=tipo_pesquisa, codigos=codigos, mensagem=mensagem)
    elif request.method == 'GET':
        return render_template('rastreamento.html', informacoes=informacoes, termo_pesquisa=termo_pesquisa, tipo_pesquisa=tipo_pesquisa, codigos=codigos)

# Define a localidade para o Brasil
locale.setlocale(locale.LC_ALL, 'pt_BR.UTF-8')

# Carregar os dados do arquivo Excel uma vez
df = pd.read_excel('Controle de Vale Pedágio - Riscos_0_1.xlsx', sheet_name='Conhecimentos Emitidos')


# Função para formatar os valores
def formatar_valor(valor):
    if isinstance(valor, str):
        # Se o valor já for uma string, retorna sem modificação
        return valor
    else:
        # Converte o valor para string e substitui os separadores de milhares e decimais
        valor_formatado = f'{valor:,.2f}'
        valor_formatado = valor_formatado.replace('.', '|').replace(',', '.').replace('|', ',')
        # Adiciona o símbolo da moeda
        return f'R$ {valor_formatado}'

# Rota para o perfil, que será a rota principal '/'
@app.route('/')
def profile():
    return render_template('profile.html')

# Rota para o mapa de calor
@app.route('/mapa_meli_arrow')
def mapa():
    return render_template('mapa_entregas.html')

# Rota para a página inicial do dashboard
@app.route('/index', methods=['GET', 'POST'])
def index():
    global df  # Acesso ao DataFrame global

    # Se a solicitação for POST, processar os filtros e retornar os dados filtrados
    if request.method == 'POST':
        # Obter os valores dos filtros do formulário
        data_inicio = request.form['data_inicio']
        data_fim = request.form['data_fim']
        filial = request.form['filial']

        # Salvar os filtros selecionados
        filtros_selecionados = {
            'data_inicio': data_inicio,
            'data_fim': data_fim,
            'filial': filial
        }

        # Aplicar filtros aos dados do DataFrame
        df_filtrado = df.copy()  # Copiar o DataFrame para preservar os dados originais
        if data_inicio and data_fim:
            df_filtrado = df_filtrado[(df_filtrado['Dt. Emissão'] >= data_inicio) & (df_filtrado['Dt. Emissão'] <= data_fim)]
        if filial:
            # Converter filial para inteiro antes de aplicar o filtro
            df_filtrado = df_filtrado[df_filtrado['Filial Cod'] == (filial)]
    else:
        # Se a solicitação for GET (primeira carga da página ou atualização), limpar os filtros e mostrar todos os dados
        df_filtrado = df.copy()
        filtros_selecionados = {
            'data_inicio': '',
            'data_fim': '',
            'filial': ''
        }

    # Calcular total de pedágio pago e formatar para moeda brasileira
    total_pedagio_pago = formatar_valor(df_filtrado['Pedagio'].sum())

    # Filiais disponíveis para seleção (incluindo todas as filiais, independentemente do filtro)
    filiais = df['Filial Cod'].unique()

    # Contagem de viagens por status atual
    viagens_por_status = df_filtrado.drop_duplicates('file editado')['Status Atual'].value_counts()

    # Contagem de viagens por tipo de frete
    viagens_por_tipo_frete = df['Tipo Frete'].value_counts()

    # Consolidar os dados por tomador de serviço e calcular a média dos riscos
    df_filtrado['Primeiro Nome'] = df_filtrado['Tomador Servico'].str.split().str[0]  # Extrair primeiro nome
    df_filtrado['Primeiro Nome'] = df_filtrado['Primeiro Nome'].str.normalize('NFKD').str.encode('ascii', errors='ignore').str.decode('utf-8')  # Remover caracteres especiais
    risco_por_tomador = df_filtrado.groupby('Primeiro Nome')['Risco'].sum().sort_values(ascending=False).map(formatar_valor)

    # Risco atual e risco reduzido
    risco_atual = formatar_valor(df_filtrado['Risco'].sum())
    risco_reduzido = formatar_valor(df_filtrado['Risco Reduzido'].sum())

    # Renderizar o template com os dados
    return render_template('index.html', total_pedagio_pago=total_pedagio_pago,
                           viagens_por_status=viagens_por_status,
                           viagens_por_tipo_frete=viagens_por_tipo_frete,
                           risco_por_tomador=risco_por_tomador,
                           risco_atual=risco_atual,
                           risco_reduzido=risco_reduzido,
                           filiais=filiais,
                           filtros=filtros_selecionados)

# Rota para a página de quantidade de viagens por status
@app.route('/viagens_por_status', methods=['GET', 'POST'])
def viagens_por_status():
    global df  # Acesso ao DataFrame global

    # Se a solicitação for POST, processar os filtros e retornar os dados filtrados
    if request.method == 'POST':
        # Obter os valores dos filtros do formulário
        data_inicio = request.form['data_inicio']
        data_fim = request.form['data_fim']
        filial = request.form['filial']

        # Salvar os filtros selecionados
        filtros_selecionados = {
            'data_inicio': data_inicio,
            'data_fim': data_fim,
            'filial': filial
        }

        # Aplicar filtros aos dados do DataFrame
        df_filtrado = df.copy()  # Copiar o DataFrame para preservar os dados originais
        if data_inicio and data_fim:
            df_filtrado = df_filtrado[(df_filtrado['Dt. Emissão'] >= data_inicio) & (df_filtrado['Dt. Emissão'] <= data_fim)]
        if filial:
            # Converter filial para inteiro antes de aplicar o filtro
            df_filtrado = df_filtrado[df_filtrado['Filial Cod'] == (filial)]
    else:
        # Se a solicitação for GET (primeira carga da página ou atualização), limpar os filtros e mostrar todos os dados
        df_filtrado = df.copy()
        filtros_selecionados = {
            'data_inicio': '',
            'data_fim': '',
            'filial': ''
        }

    # Contagem de viagens por status atual
    viagens_por_status = df_filtrado.drop_duplicates('file editado')['Status Atual'].value_counts()

    # Filiais disponíveis para seleção (incluindo todas as filiais, independentemente do filtro)
    filiais = df['Filial Cod'].unique()

    # Contagem de viagens por status atual e emissor
    # Converter todos os nomes de usuário para minúsculas antes de agrupar
    df_filtrado['Usuário Emissor'] = df_filtrado['Usuário Emissor'].str.lower()
    viagens_por_status_e_emissor = df_filtrado.drop_duplicates(['file editado', 'Usuário Emissor']) \
                                     .groupby(['Usuário Emissor', 'Status Atual']).size().unstack(fill_value=0)

    # Converter os dados para HTML
    viagens_por_status_e_emissor_html = viagens_por_status_e_emissor.to_html(classes='data-table custom-table', header='true')

    # Renderizar o template com os dados
    return render_template('viagens_por_status.html', viagens_por_status=viagens_por_status,
                           viagens_por_status_e_emissor_html=viagens_por_status_e_emissor_html,
                           filiais=filiais, filtros=filtros_selecionados)


# Rota para a tabela de dados filtrados
@app.route('/tabela_filtrada', methods=['POST'])
def tabela_filtrada():
    global df  # Acesso ao DataFrame global

    # Obter os valores dos filtros do formulário
    data_inicio = request.form['data_inicio']
    data_fim = request.form['data_fim']
    filial = request.form['filial']

    # Aplicar filtros aos dados do DataFrame
    df_filtrado = df.copy()  # Copiar o DataFrame para preservar os dados originais
    if data_inicio and data_fim:
        df_filtrado = df_filtrado[(df_filtrado['Dt. Emissão'] >= data_inicio) & (df_filtrado['Dt. Emissão'] <= data_fim)]
    if filial:
        # Converter filial para inteiro antes de aplicar o filtro
        df_filtrado = df_filtrado[df_filtrado['Filial Cod'] == (filial)]

    # Retornar os dados filtrados em formato JSON
    return df_filtrado.to_json(orient='records')

def buscar_noticias_logistica():
    url = "http://newsapi.org/v2/top-headlines?country=br&apiKey=332b5d3d54214270ba2882d7245c410f"
    response = requests.get(url)
    noticias = response.json()['articles']
    return noticias

@app.route('/noticias')
def noticias():
    noticias = buscar_noticias_logistica()
    return render_template('noticias.html', noticias=noticias)

if __name__ == '__main__':
    app.run(debug=True)
